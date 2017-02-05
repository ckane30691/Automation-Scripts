import openpyxl
import time
import reportlab.platypus
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


def import_excel_data():
	distro_file_name = raw_input("Please enter distro file name: ")
	distrowb = openpyxl.load_workbook(distro_file_name)
	distrowbArray = distrowb.get_sheet_names()
	distrosheet = distrowb.get_sheet_by_name(distrowbArray[0])
	jobNumber = raw_input("Please enter job #: ")
	poNumber = raw_input("Please enter PO #: ")
	item_desc_array = []
	packing_slip_name = jobNumber + " Commerical Invoice.pdf"
	c = canvas.Canvas(packing_slip_name, pagesize=letter)
	#COLUMN START POINT NEEDS TO BE ADJUSTED TO TEST WITH ACTUAL DISTRO LIST
	for y in range(9, distrosheet.max_column + 1, 1):
		if y % 2 != 0:
			item_desc = distrosheet.cell(row=1, column=y).value
			item_desc_array.append(item_desc)
	for x in range(2, distrosheet.max_row + 1, 1):
		shipToName = distrosheet.cell(row=x, column=1).value
		address1 = distrosheet.cell(row=x, column=2).value
		address2 = distrosheet.cell(row=x, column=3).value
		if isinstance(address2, long) == False and address2 != None:
			address2 = address2.encode('ascii', 'ignore').decode('ascii')
		else:
			address2 = str(address2)
		city = distrosheet.cell(row=x, column=4).value
		state = distrosheet.cell(row=x, column=5).value
		zipcode = distrosheet.cell(row=x, column=6).value
		if isinstance(zipcode, long) == False and zipcode != None:
			zipcode = zipcode.encode('ascii', 'ignore').decode('ascii')
		else:
			zipcode = str(zipcode)
		country = distrosheet.cell(row=x, column=7).value
		qtyOrderedArray = []
		itemCostArray = []
		#COLUMN START POINT NEEDS TO BE ADJUSTED TO TEST WITH ACTUAL DISTRO LIST
		for q in range(9, distrosheet.max_column + 1, 1):
			if q % 2 != 0:
				itemQtyOrdered = distrosheet.cell(row=x, column=q).value
				qtyOrderedArray.append(str(itemQtyOrdered))
			else:
				itemCost = distrosheet.cell(row=x, column=q).value
				itemCostArray.append(itemCost)
		generate_packing_slip(c, shipToName, address1, address2, city, state, zipcode, country, jobNumber, item_desc_array, qtyOrderedArray, itemCostArray, poNumber)
	c.save()	


def generate_packing_slip(c, shipToName, address1, address2, city, state, zipcode, country, jobNumber, item_desc_array, qtyOrderedArray, itemCostArray, poNumber):
	background = "CI_template.jpg"
	c.drawImage(background, 0, 0, width=612, height=792)
	#Ship/Sold to Address
	c.drawString(94, 435, shipToName)
	c.drawString(94, 423, address1)
	if city == None:
		city = ''
	if state == None:
		state = ''
	if zipcode == None:
		zipcode = ''
	if zipcode == "None":
		zipcode = ''
	if country == None:
		country = ''
	if address2 == "None":
		c.drawString(94, 410, city + ', ' + state + ' ' + zipcode + ' ' + country)
	else:
		c.drawString(94, 410, address2)
		c.drawString(94, 397, city + ', ' + state + ' ' + zipcode + ' ' + country)

	#Job Number
	c.drawString(180, 500, jobNumber)

	#PO Number
	c.drawString(137, 348, poNumber)

	#Ship Date
	c.drawString(485, 500, time.strftime("%m/%d/%Y"))

	#Item Description & QTYs
	total = 0
	count = 310
	for z in range(0, len(item_desc_array), 1):
		if qtyOrderedArray[z] != "None" and qtyOrderedArray[z] != "0":
			extension = float(itemCostArray[z]) * float(qtyOrderedArray[z])
			c.drawString(120, count, item_desc_array[z])
			c.drawString(370, count, qtyOrderedArray[z])
			c.drawString(420, count, "$" + str(itemCostArray[z]))
			c.drawString(520, count, "$" + str(extension))
			total += extension
			count -= 10
	c.drawString(520, 217, "$" + str(total))
	c.showPage()
	


import_excel_data()